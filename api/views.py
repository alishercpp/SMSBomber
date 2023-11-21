from django.shortcuts import render
from openpyxl import Workbook, load_workbook
from rest_framework.decorators import api_view
from rest_framework.response import Response
from datetime import datetime, timedelta
from .models import User, Device
import requests


def home(request):
    return render(request, "home.html")


@api_view(http_method_names=['GET', 'POST'])
def logout(request):
    if request.method == "POST":
        phone = request.data.get("phone_number")
        uid = request.data.get("uid")
        user = User.objects.filter(username=phone)
        if user:
            device = Device.objects.filter(uid=uid)
            if device:
                device.first().delete()
                return Response({
                    "status": "logout",
                    "days": 0,
                })
        return Response({
            "status": "false",
            "days": 0,
        })
    return Response({
        "status": "false"
    })
# test
@api_view(http_method_names=['GET', 'POST'])
def login(request):
    if request.method == "POST":
        username = request.data.get("phone_number")
        password = request.data.get("password")
        hour = request.data.get("hour")
        uid = request.data.get("uid")
        state = request.data.get("state")
        if int(hour) < 7 or int(hour) > 19:
            return Response({
                "status": "timeout",
            })
        user = User.objects.filter(username=username)
        if user:
            user = user.first()
            now = datetime.now()
            days = (user.end_date.date() - now.date()).days
            device = Device.objects.filter(uid=uid)
            if device and state == "login":
                return Response({
                    "status": "false",
                    "days": -1
                })
            if int(days) < 0:
                return Response({
                    "status": "stopped",
                    "days": -2,
                })
            if not user.is_free:
                return Response({
                    "status": "false", 
                    "days": -3
                })
            if password == user.token:
                Device.objects.create(
                    user=user,
                    uid=uid
                )
                return Response({
                    "status": "true",
                    "days": days,
                })
            return Response({
                "status": "false",
                "days": -4
            })
        print(2)
    print(1)
    return Response({
        "status": "false",
    })

def check(number):
    if number.strip().isdigit() or "+" in number:
        return True
    return False

def update(number):
    number = number.strip()
    if len(number) == 12:
        number = "+" + number
    elif len(number) == 9:
        number = "+998" + number
    return number


@api_view(http_method_names=['GET', 'POST'])
def parse_excel(request):
    if request.method == "POST":
        username = request.data.get("phone_number")
        password = request.data.get("password")
        file = request.data.get("file")
        hour = request.data.get("hour")
        if int(hour) < 7 or int(hour) > 19:
            return Response({
                "status": "timeout",
            })
        user = User.objects.filter(username=username)
        if user:
            user = user.first()
            if not user.is_free:
                return Response({
                    "status": "false",  
                })
            if password == user.token:
                now = datetime.now()
                days = (user.end_date.date() - now.date()).days
                if int(days) < 0:
                    return Response({
                        "status": "stopped",
                    })
                try:
                    res = requests.get(url=file)
                    with open(f"{user.token}.xlsx", "wb") as f:
                        f.write(res.content)
                    wb = load_workbook(f"{user.token}.xlsx")
                    # wb = load_workbook(f"new.xlsx")
                    ws = wb.active
                    r = {}
                    r["data"] = []
                    r["days"] = days
                    c = 0
                    for row in ws.iter_rows():
                        text = []
                        data = {}
                        key = None
                        for cell in row:
                            if cell.value == None:
                                continue
                            if cell.column == 1:
                                key = cell.value
                                continue
                            text.append(cell.value)
                        if text:
                            data["phones"] = list(map(update, list(filter(check, str(key).split(",")))))
                            messages = " ".join(str(i) for i in text)
                            data["message"] = messages
                        if data:
                            r["data"].append(data)
                        text = []
                        key = None
                        data = {}
                        c += 1
                    return Response(r)
                except Exception as e:
                    print(e)
                    return Response({
                        "status": "false",
                    })
            return Response({
                "status": "false",
            })
    return Response({
        "status": "false"
    })